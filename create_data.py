import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

np.random.seed(42)

# ── SUBURB DATA ───────────────────────────────────────
suburbs_data = [
    # suburb, lat, lon, type, prestige_factor
    ("Melbourne CBD", -37.8136, 144.9631, "Inner", 1.45),
    ("Southbank", -37.8230, 144.9680, "Inner", 1.35),
    ("Docklands", -37.8145, 144.9440, "Inner", 1.25),
    ("Carlton", -37.7990, 144.9670, "Inner", 1.20),
    ("Fitzroy", -37.7990, 144.9780, "Inner", 1.22),
    ("Collingwood", -37.8040, 144.9860, "Inner", 1.15),
    ("Richmond", -37.8180, 144.9990, "Inner", 1.18),
    ("South Yarra", -37.8390, 144.9930, "Inner", 1.45),
    ("Toorak", -37.8440, 145.0110, "Inner", 1.80),
    ("Prahran", -37.8490, 144.9890, "Inner", 1.30),
    ("St Kilda", -37.8678, 144.9814, "Inner", 1.25),
    ("Port Melbourne", -37.8350, 144.9330, "Inner", 1.30),
    ("Albert Park", -37.8430, 144.9560, "Inner", 1.40),
    ("Middle Park", -37.8510, 144.9560, "Inner", 1.38),
    ("Hawthorn", -37.8220, 145.0330, "Middle", 1.35),
    ("Kew", -37.8080, 145.0330, "Middle", 1.30),
    ("Camberwell", -37.8420, 145.0580, "Middle", 1.28),
    ("Brighton", -37.9050, 144.9990, "Middle", 1.55),
    ("Elwood", -37.8820, 144.9860, "Middle", 1.35),
    ("Caulfield", -37.8780, 145.0220, "Middle", 1.20),
    ("Northcote", -37.7720, 145.0010, "Middle", 1.15),
    ("Preston", -37.7470, 145.0070, "Middle", 1.00),
    ("Thornbury", -37.7610, 144.9990, "Middle", 1.10),
    ("Brunswick", -37.7660, 144.9600, "Middle", 1.18),
    ("Coburg", -37.7440, 144.9650, "Middle", 1.05),
    ("Box Hill", -37.8190, 145.1220, "Middle", 1.08),
    ("Glen Waverley", -37.8780, 145.1640, "Middle", 1.12),
    ("Footscray", -37.8010, 144.8990, "Middle", 1.00),
    ("Williamstown", -37.8610, 144.8990, "Middle", 1.18),
    ("Newport", -37.8460, 144.8820, "Middle", 1.10),
    ("Sunshine", -37.7880, 144.8310, "Outer", 0.82),
    ("Altona", -37.8700, 144.8280, "Outer", 0.88),
    ("Werribee", -37.9040, 144.6620, "Outer", 0.72),
    ("Dandenong", -37.9870, 145.2150, "Outer", 0.75),
    ("Frankston", -38.1440, 145.1260, "Outer", 0.78),
    ("Ringwood", -37.8140, 145.2280, "Outer", 0.88),
    ("Croydon", -37.7960, 145.2830, "Outer", 0.82),
    ("Knox", -37.8710, 145.2420, "Outer", 0.85),
    ("Berwick", -38.0340, 145.3560, "Outer", 0.80),
    ("Pakenham", -38.0710, 145.4870, "Outer", 0.68),
]

rows = []
for suburb, lat, lon, stype, pf in suburbs_data:
    dist = np.sqrt((lat-(-37.8136))**2 + (lon-144.9631)**2) * 111

    r1 = max(250, int(400*pf + np.random.normal(0,15)))
    r2 = max(350, int(580*pf + np.random.normal(0,20)))
    r3 = max(450, int(780*pf + np.random.normal(0,25)))

    transport = max(35, min(99, int(92 - dist*3.2 + np.random.normal(0,6))))
    safety    = max(45, min(99, int(72 + np.random.normal(0,10))))
    walk      = max(30, min(99, int(88 - dist*2.8 + np.random.normal(0,8))))
    amenities = max(30, min(99, int(85 - dist*2.2 + np.random.normal(0,10))))

    grocery   = max(80,  int(110 + dist*1.8 + np.random.normal(0,8)))
    transport_cost = max(40, int(55 + dist*3.2 + np.random.normal(0,6)))
    dining    = max(80,  int(180*pf + np.random.normal(0,15)))
    utilities = max(50,  int(80 + np.random.normal(0,10)))

    total = r1 + grocery + transport_cost + int(dining*0.35) + utilities

    rows.append({
        "Suburb": suburb,
        "Latitude": round(lat, 4),
        "Longitude": round(lon, 4),
        "Suburb_Type": stype,
        "Distance_to_CBD_km": round(dist, 1),
        "Rent_1BR_per_week": r1,
        "Rent_2BR_per_week": r2,
        "Rent_3BR_per_week": r3,
        "Transport_Score_out_of_100": transport,
        "Safety_Score_out_of_100": safety,
        "Walkability_Score_out_of_100": walk,
        "Amenities_Score_out_of_100": amenities,
        "Weekly_Groceries_AUD": grocery,
        "Weekly_Transport_Cost_AUD": transport_cost,
        "Weekly_Dining_Budget_AUD": dining,
        "Weekly_Utilities_AUD": utilities,
        "Weekly_Total_Living_Cost_AUD": total,
        "Data_Source": "Simulated — Replace with real data",
        "Last_Updated": "2026-Q1"
    })

df_suburbs = pd.DataFrame(rows)

# ── RENT TRENDS DATA ──────────────────────────────────
trend_suburbs = [
    "Melbourne CBD", "South Yarra", "Brunswick",
    "Footscray", "Brighton", "Dandenong",
    "Toorak", "St Kilda", "Sunshine", "Frankston"
]
base_rents = {
    "Melbourne CBD": 480, "South Yarra": 560, "Brunswick": 390,
    "Footscray": 330, "Brighton": 620, "Dandenong": 290,
    "Toorak": 750, "St Kilda": 450, "Sunshine": 310, "Frankston": 310
}
quarters = pd.date_range(start='2019-01-01', end='2024-12-31', freq='QS')
trend_rows = []
for suburb in trend_suburbs:
    base = base_rents[suburb]
    for i, q in enumerate(quarters):
        covid = -35 if str(q.year) in ['2020'] else (-15 if '2021-01' in str(q) else 0)
        rent = max(200, int(base + i*4.2 + covid + np.random.normal(0,7)))
        trend_rows.append({
            "Suburb": suburb,
            "Quarter": f"{q.year}-Q{(q.month-1)//3+1}",
            "Date": q.strftime("%Y-%m-%d"),
            "Median_Weekly_Rent_AUD": rent,
            "YoY_Change_Percent": round(np.random.normal(4.5, 2.5), 1),
            "Data_Source": "Simulated — Replace with real data"
        })
df_trends = pd.DataFrame(trend_rows)

# ── COST OF LIVING DATA ───────────────────────────────
cost_rows = []
categories = [
    ("1BR Rent", "Housing"),
    ("2BR Rent", "Housing"),
    ("3BR Rent", "Housing"),
    ("Weekly Groceries", "Food"),
    ("Weekly Transport", "Transport"),
    ("Weekly Dining", "Lifestyle"),
    ("Weekly Utilities", "Utilities"),
    ("Weekly Total (1BR)", "Total"),
]
for suburb, lat, lon, stype, pf in suburbs_data:
    row = df_suburbs[df_suburbs["Suburb"]==suburb].iloc[0]
    cost_rows.append({
        "Suburb": suburb,
        "Suburb_Type": stype,
        "Rent_1BR": row["Rent_1BR_per_week"],
        "Rent_2BR": row["Rent_2BR_per_week"],
        "Rent_3BR": row["Rent_3BR_per_week"],
        "Groceries_pw": row["Weekly_Groceries_AUD"],
        "Transport_pw": row["Weekly_Transport_Cost_AUD"],
        "Dining_pw": row["Weekly_Dining_Budget_AUD"],
        "Utilities_pw": row["Weekly_Utilities_AUD"],
        "Total_Weekly_1BR": row["Weekly_Total_Living_Cost_AUD"],
        "Affordability_Index": round(100 - (row["Rent_1BR_per_week"]/10), 1),
        "Data_Source": "Simulated — Replace with real data"
    })
df_cost = pd.DataFrame(cost_rows)

# ── WRITE TO EXCEL ────────────────────────────────────
wb = Workbook()

# Style helpers
def header_style(ws, row_num, color="1F4E79"):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        bottom=Side(style="medium", color="FFFFFF")
    )
    for cell in ws[row_num]:
        if cell.value:
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

def add_sheet(wb, df, sheet_name, header_color):
    ws = wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    header_style(ws, 1, header_color)
    # Alternate row colors
    light = PatternFill("solid", fgColor="F0F4FA")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                cell.fill = light
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    auto_width(ws)
    ws.freeze_panes = "A2"
    return ws

# Remove default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# Sheet 1 — Suburb Overview
ws1 = add_sheet(wb, df_suburbs, "Suburb_Overview", "1F4E79")
ws1.sheet_properties.tabColor = "1F4E79"

# Sheet 2 — Rent Trends
ws2 = add_sheet(wb, df_trends, "Rent_Trends", "375623")
ws2.sheet_properties.tabColor = "375623"

# Sheet 3 — Cost of Living
ws3 = add_sheet(wb, df_cost, "Cost_of_Living", "7B2C8B")
ws3.sheet_properties.tabColor = "7B2C8B"

# Sheet 4 — Data Dictionary
ws4 = wb.create_sheet(title="Data_Dictionary")
ws4.sheet_properties.tabColor = "C44FD4"
dict_data = [
    ["Column Name", "Description", "Unit", "Sheet"],
    ["Suburb", "Name of Melbourne suburb", "Text", "All"],
    ["Latitude", "Geographic latitude coordinate", "Decimal degrees", "Suburb_Overview"],
    ["Longitude", "Geographic longitude coordinate", "Decimal degrees", "Suburb_Overview"],
    ["Suburb_Type", "Inner (<5km), Middle (5-15km), Outer (>15km)", "Category", "Suburb_Overview"],
    ["Distance_to_CBD_km", "Distance from Melbourne CBD", "Kilometres", "Suburb_Overview"],
    ["Rent_1BR_per_week", "Median weekly rent for 1 bedroom unit", "AUD $", "Suburb_Overview"],
    ["Rent_2BR_per_week", "Median weekly rent for 2 bedroom unit", "AUD $", "Suburb_Overview"],
    ["Rent_3BR_per_week", "Median weekly rent for 3 bedroom unit", "AUD $", "Suburb_Overview"],
    ["Transport_Score_out_of_100", "Public transport accessibility score", "Score 0-100", "Suburb_Overview"],
    ["Safety_Score_out_of_100", "Suburb safety rating", "Score 0-100", "Suburb_Overview"],
    ["Walkability_Score_out_of_100", "Pedestrian accessibility score", "Score 0-100", "Suburb_Overview"],
    ["Amenities_Score_out_of_100", "Access to shops, parks, schools", "Score 0-100", "Suburb_Overview"],
    ["Weekly_Groceries_AUD", "Estimated weekly grocery spend", "AUD $", "Suburb_Overview"],
    ["Weekly_Transport_Cost_AUD", "Estimated weekly public transport cost", "AUD $", "Suburb_Overview"],
    ["Weekly_Dining_Budget_AUD", "Estimated weekly dining out budget", "AUD $", "Suburb_Overview"],
    ["Weekly_Utilities_AUD", "Estimated weekly utilities (electricity, gas, internet)", "AUD $", "Suburb_Overview"],
    ["Weekly_Total_Living_Cost_AUD", "Total estimated weekly cost of living (1BR)", "AUD $", "Suburb_Overview"],
    ["Quarter", "Financial quarter (e.g. 2024-Q1)", "Text", "Rent_Trends"],
    ["Median_Weekly_Rent_AUD", "Median weekly rent for that quarter", "AUD $", "Rent_Trends"],
    ["YoY_Change_Percent", "Year-on-year rent change percentage", "Percent %", "Rent_Trends"],
    ["Affordability_Index", "Higher = more affordable (100 - rent/10)", "Index", "Cost_of_Living"],
    ["Data_Source", "Source of data — replace Simulated with real source", "Text", "All"],
    ["Last_Updated", "When data was last updated", "Text", "Suburb_Overview"],
]
for row in dict_data:
    ws4.append(row)
header_style(ws4, 1, "C44FD4")
for i, row in enumerate(ws4.iter_rows(min_row=2), start=2):
    if i % 2 == 0:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="F9F0FF")
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
auto_width(ws4)
ws4.freeze_panes = "A2"

# Save
wb.save("melbourne_housing_data.xlsx")
print("✅ Excel file created: melbourne_housing_data.xlsx")
print(f"   Sheet 1 — Suburb_Overview: {len(df_suburbs)} suburbs")
print(f"   Sheet 2 — Rent_Trends: {len(df_trends)} quarterly records")
print(f"   Sheet 3 — Cost_of_Living: {len(df_cost)} suburb cost breakdowns")
print(f"   Sheet 4 — Data_Dictionary: column descriptions")
print("\n📌 To use real data: replace values in Excel keeping same column names!")